from datetime import datetime
from io import BytesIO
import pandas as pd
import streamlit as st
from streamlit_gsheets import GSheetsConnection

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ==========================================
# 1. PAGE CONFIGURATION & CUSTOM CSS
# ==========================================
st.set_page_config(page_title="District Book Inventory", layout="wide")

st.markdown("""
    <style>
    /* Dark Slate Sidebar styling */
    [data-testid="stSidebar"] {
        background-color: #1E293B !important;
    }
    [data-testid="stSidebar"] * {
        color: #FFFFFF !important;
    }

    /* Fix Select Box Container Background & Borders */
    div[data-baseweb="select"] > div {
        background-color: #1E293B !important;
        border: 1px solid #475569 !important;
    }

    /* Force Select Box Selected Text & Arrow Icon to White */
    div[data-baseweb="select"] span,
    div[data-baseweb="select"] div,
    div[data-baseweb="select"] svg {
        color: #FFFFFF !important;
        fill: #FFFFFF !important;
    }

    /* Fix Main Page Input Fields */
    textarea, div[data-baseweb="input"] > div {
        background-color: #FFFFFF !important;
        color: #0F172A !important;
        border: 1px solid #CBD5E1 !important;
    }

    input, textarea {
        color: #0F172A !important;
    }

    .stTextInput label, .stDateInput label, .stTextArea label, .stSelectbox label {
        color: #0F172A !important;
        font-weight: 600;
    }
    </style>
""", unsafe_allow_html=True)

st.title("📚 District Book Inventory Tracker")

# --- Top Metrics Cards ---
col1, col2, col3 = st.columns(3)
with col1:
    st.metric(label="🏫 Total Schools", value="16")
with col2:
    st.metric(label="📦 Pending Dispatches", value="12")
with col3:
    st.metric(label="✅ Completed Orders", value="148")

st.divider()

# ==========================================
# 2. CONNECT TO GOOGLE SHEETS & PRESETS
# ==========================================
conn = st.connection("gsheets", type=GSheetsConnection)

def load_data(worksheet_name):
    return conn.read(worksheet=worksheet_name, ttl=0)

SCHOOL_LIST = [
    "Arcaflor Maniapao ES",
    "Balabag ES",
    "Casildo B. Nonol Sr. ES",
    "Colorado ES",
    "Damñas ES",
    "Digos City Central ES",
    "Domingo Abawag ES",
    "Dulangan ES",
    "Federico Alferez ES",
    "Jolencio R. Alberca ES",
    "Lungag ES",
    "Mahayahay ES",
    "Pedro Basalan ES",
    "Ranao ES",
    "Remedios N. Saplala ES",
    "Ruparan ES",
]

role = st.sidebar.radio("Select View:", ["Principal View", "Custodian View"])

# ==========================================
# 3. FORMAL DEPED ICS EXCEL GENERATOR
# ==========================================
def generate_official_deped_ics_excel(school_name, date_str, df_items):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ICS Form"

    # Ensure Gridlines are visible on print & view
    ws.views.sheetView[0].showGridLines = True

    # Styling Assets
    font_header = Font(name="Calibri", size=10)
    font_title = Font(name="Calibri", size=14, bold=True)
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_regular = Font(name="Calibri", size=10)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    thin_border_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # 1. Header (Rows 1-5)
    headers = [
        "Republic of the Philippines",
        "Department of Education",
        "Region XI – Davao Region",
        "City Schools Division of Digos",
        "Digos Occidental District"
    ]
    for r_idx, text in enumerate(headers, start=1):
        ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=7)
        cell = ws.cell(row=r_idx, column=1, value=text)
        cell.font = font_header
        cell.alignment = align_center

    # 2. Section Title (Row 7)
    ws.merge_cells("A7:G7")
    title_cell = ws.cell(row=7, column=1, value="INVENTORY CUSTODIAN SLIP (ICS)")
    title_cell.font = font_title
    title_cell.alignment = align_center

    # 3. Metadata Section (Rows 11-12)
    ws.cell(row=11, column=1, value=f"Entity Name: {school_name}").font = font_bold
    ws.cell(row=12, column=1, value="INVENTORY CUSTODIAN SLIP (ICS) No.: ____________________").font = font_bold

    # 4. Table Column Headers (Row 14)
    table_headers = ["Quantity", "Unit", "Unit Cost", "Total Cost", "Description", "Inventory Item No.", "Estimated Useful Life"]
    ws.row_dimensions[14].height = 28
    
    for c_idx, h_text in enumerate(table_headers, start=1):
        cell = ws.cell(row=14, column=c_idx, value=h_text)
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = thin_border

    # 5. Populate Data Rows (Starting at Row 15)
    current_row = 15
    for _, row in df_items.iterrows():
        qty = int(row.get("quantity_received", 0))
        unit_cost = float(row.get("unit_cost", 90.00)) if "unit_cost" in row else 90.00
        desc = str(row.get("book_title", ""))
        useful_life = int(row.get("useful_life", 3))

        ws.cell(row=current_row, column=1, value=qty).alignment = align_center
        ws.cell(row=current_row, column=2, value="PCS").alignment = align_center
        
        c3 = ws.cell(row=current_row, column=3, value=unit_cost)
        c3.number_format = '#,##0.00'
        c3.alignment = align_right

        # Excel formula for Total Cost (= Quantity * Unit Cost)
        c4 = ws.cell(row=current_row, column=4, value=f"=A{current_row}*C{current_row}")
        c4.number_format = '#,##0.00'
        c4.alignment = align_right

        ws.cell(row=current_row, column=5, value=desc).alignment = align_left
        ws.cell(row=current_row, column=6, value="").alignment = align_center
        ws.cell(row=current_row, column=7, value=useful_life).alignment = align_center

        for col in range(1, 8):
            ws.cell(row=current_row, column=col).border = thin_border
            ws.cell(row=current_row, column=col).font = font_regular

        current_row += 1

    # Fill blank padded rows to keep standard sheet length
    target_end_row = max(current_row + 3, 26)
    for r in range(current_row, target_end_row):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c, value="")
            cell.border = thin_border

    # 6. Signatures Block
    sig_start = target_end_row + 1
    
    # Left Header (Custodian)
    ws.merge_cells(start_row=sig_start, start_column=1, end_row=sig_start, end_column=4)
    c_from = ws.cell(row=sig_start, column=1, value="Received from:")
    c_from.font = font_bold
    
    # Right Header (Recipient)
    ws.merge_cells(start_row=sig_start, start_column=5, end_row=sig_start, end_column=7)
    c_by = ws.cell(row=sig_start, column=5, value="Received by:")
    c_by.font = font_bold

    # Custodian Name
    ws.merge_cells(start_row=sig_start+2, start_column=1, end_row=sig_start+2, end_column=4)
    c_cust = ws.cell(row=sig_start+2, column=1, value="HERICK REEL D. SORDILLA")
    c_cust.font = font_bold
    c_cust.alignment = align_center

    ws.merge_cells(start_row=sig_start+3, start_column=1, end_row=sig_start+3, end_column=4)
    ws.cell(row=sig_start+3, column=1, value="District Property Custodian").alignment = align_center

    ws.merge_cells(start_row=sig_start+4, start_column=1, end_row=sig_start+4, end_column=4)
    ws.cell(row=sig_start+4, column=1, value="Date: ____________________").alignment = align_center

    # Recipient Line
    ws.merge_cells(start_row=sig_start+2, start_column=5, end_row=sig_start+2, end_column=7)
    ws.cell(row=sig_start+2, column=5, value="__________________________________").alignment = align_center

    ws.merge_cells(start_row=sig_start+3, start_column=5, end_row=sig_start+3, end_column=7)
    ws.cell(row=sig_start+3, column=5, value="Signature over Printed Name of End-User").alignment = align_center

    ws.merge_cells(start_row=sig_start+4, start_column=5, end_row=sig_start+4, end_column=7)
    ws.cell(row=sig_start+4, column=5, value=f"Date: {date_str}").alignment = align_center

    # Signatures Outer Border Outline
    for r in range(sig_start, sig_start+5):
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = thin_border

    # Set Column Widths for clean layout
    col_widths = {1: 12, 2: 10, 3: 14, 4: 16, 5: 32, 6: 20, 7: 20}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ==========================================
# 4. CUSTODIAN VIEW
# ==========================================
if role == "Custodian View":
    st.header("🔒 Custodian Control Panel")

    password = st.text_input("Enter Custodian Password to Access:", type="password")
    CUSTODIAN_PASSWORD = "admin123"

    if password == CUSTODIAN_PASSWORD:
        st.success("Access Granted!")

        col1, col2 = st.columns([1, 2])

        # --- A. ADD NEW BOOKS TO CENTRAL STORAGE ---
        with col1:
            st.subheader("Add / Update Central Stock")
            with st.form("add_book_form"):
                title = st.text_input("Book Title")
                stock = st.number_input("Quantity to Add", min_value=1, step=1, value=100)
                submit = st.form_submit_button("Add to Master Stock")

                if submit and title:
                    master_df = load_data("master_inventory")
                    title_clean = title.strip()

                    if not master_df.empty and title_clean in master_df["book_title"].values:
                        master_df.loc[master_df["book_title"] == title_clean, "central_stock"] += stock
                    else:
                        new_row = pd.DataFrame([{"book_title": title_clean, "central_stock": stock}])
                        master_df = pd.concat([master_df, new_row], ignore_index=True)

                    conn.update(worksheet="master_inventory", data=master_df)
                    st.success(f"Added {stock} copies of '{title_clean}'!")
                    st.rerun()

        # --- B. DISPATCH BOOKS GRID ---
        with col2:
            st.subheader("Dispatch Books to Schools")
            master_df = load_data("master_inventory")

            if master_df.empty:
                st.info("Please add books to Central Stock first before dispatching.")
            else:
                master_df = master_df.sort_values(by="book_title", ascending=False)
                book_options = master_df["book_title"].tolist()

                m_col1, m_col2 = st.columns([3, 1])
                selected_book = m_col1.selectbox("📖 Select Book Title to Dispatch:", book_options)
                dispatch_all_btn = m_col2.button("⚡ Batch Dispatch All", type="primary", use_container_width=True)

                current_stock = int(master_df.loc[master_df["book_title"] == selected_book, "central_stock"].values[0])

                st.divider()

                live_total_requested = 0
                dispatch_selections = []

                for idx, school in enumerate(SCHOOL_LIST):
                    c1, c2, c3 = st.columns([4, 2, 2])
                    c1.write(f"**{school}**")

                    input_key = f"dispatch_qty_{selected_book}_{idx}"
                    if input_key not in st.session_state:
                        st.session_state[input_key] = 10

                    qty = c2.number_input(
                        f"Qty for {school}",
                        min_value=1,
                        step=1,
                        key=input_key,
                        label_visibility="collapsed",
                    )
                    live_total_requested += qty
                    dispatch_selections.append({"school": school, "qty": qty})

                    if c3.button("Dispatch", key=f"dispatch_btn_{selected_book}_{idx}"):
                        if qty > current_stock:
                            st.error(f"Not enough stock! Only {current_stock} available.")
                        else:
                            master_df.loc[master_df["book_title"] == selected_book, "central_stock"] -= qty
                            conn.update(worksheet="master_inventory", data=master_df)

                            school_df = load_data("school_inventory")
                            new_dispatch = pd.DataFrame([{
                                "school_name": school,
                                "book_title": selected_book,
                                "quantity_received": qty,
                                "status": "For Release"
                            }])
                            school_df = pd.concat([school_df, new_dispatch], ignore_index=True)
                            conn.update(worksheet="school_inventory", data=school_df)

                            st.success(f"Dispatched {qty} copies of '{selected_book}' to {school}!")
                            st.rerun()

                st.divider()
                m1, m2, m3 = st.columns(3)
                m1.metric("📦 Warehouse Stock Available", current_stock)
                m2.metric("📋 Total Input (All Schools)", live_total_requested)
                m3.metric("🟢 Remaining Stock After Batch", current_stock - live_total_requested)

                if dispatch_all_btn:
                    if live_total_requested > current_stock:
                        st.error(f"Cannot batch dispatch! Required `{live_total_requested}` copies, but only `{current_stock}` available.")
                    else:
                        master_df.loc[master_df["book_title"] == selected_book, "central_stock"] -= live_total_requested
                        conn.update(worksheet="master_inventory", data=master_df)

                        school_df = load_data("school_inventory")
                        new_rows = pd.DataFrame([
                            {
                                "school_name": item["school"],
                                "book_title": selected_book,
                                "quantity_received": item["qty"],
                                "status": "For Release"
                            } for item in dispatch_selections
                        ])
                        school_df = pd.concat([school_df, new_rows], ignore_index=True)
                        conn.update(worksheet="school_inventory", data=school_df)

                        st.success(f"Successfully batch dispatched **{selected_book}** to all schools!")
                        st.rerun()

        st.divider()

        # --- C. DATA & MANAGEMENT TABS ---
        tab1, tab2, tab3 = st.tabs(["📊 Central Warehouse Stock", "🚚 Dispatched Inventory Log", "📅 Scheduled Appointments"])

        with tab1:
            central_df = load_data("master_inventory")
            if not central_df.empty:
                st.dataframe(central_df.sort_values(by="book_title", ascending=False), use_container_width=True)

        with tab2:
            dispatched_df = load_data("school_inventory")
            if not dispatched_df.empty:
                st.dataframe(dispatched_df.sort_values(by="book_title", ascending=False), use_container_width=True)

        with tab3:
            appointments_df = load_data("appointments")
            if not appointments_df.empty:
                st.dataframe(appointments_df, use_container_width=True)

# ==========================================
# 5. PRINCIPAL VIEW
# ==========================================
else:
    st.header("Principal Portal")
    selected_school = st.selectbox("Select Your School:", SCHOOL_LIST)

    # Notice Box
    st.info(f"ℹ️ **Notice for {selected_school}:** Please be informed that the following books/learning materials assigned to your school are now ready for pickup at the District Office.")

    school_df = load_data("school_inventory")
    master_df = load_data("master_inventory")

    if not school_df.empty:
        filtered = school_df[school_df["school_name"].astype(str).str.strip() == selected_school.strip()]
        
        # Filter out items already marked as Received
        if not filtered.empty:
            filtered = filtered[filtered["status"].astype(str).str.strip().str.lower() != "received"]
        
        if not filtered.empty:
            summary = filtered.groupby(["book_title", "status"])["quantity_received"].sum().reset_index()
            
            # --- 🔗 MERGE BOOK PRICES & USEFUL LIFE FROM MASTER INVENTORY ---
            if not master_df.empty and "unit_cost" in master_df.columns:
                # Merge unit_cost and useful_life from master inventory based on book_title
                cols_to_merge = ["book_title", "unit_cost"]
                if "useful_life" in master_df.columns:
                    cols_to_merge.append("useful_life")
                
                summary = summary.merge(master_df[cols_to_merge], on="book_title", how="left")
            else:
                # Fallback defaults if unit_cost column doesn't exist yet in Google Sheets
                summary["unit_cost"] = 90.00
                summary["useful_life"] = 3

            # Fill any missing values if a book title wasn't found in master
            summary["unit_cost"] = summary["unit_cost"].fillna(90.00)
            summary["useful_life"] = summary["useful_life"].fillna(3)

            summary = summary.sort_values(by="book_title", ascending=False)
            
            # Display readable summary on web interface
            st.dataframe(summary[["book_title", "status", "quantity_received", "unit_cost"]], use_container_width=True)

            # --- 📊 GENERATE & DOWNLOAD OFFICIAL DEPED ICS EXCEL ---
            excel_data = generate_official_deped_ics_excel(
                school_name=selected_school,
                date_str=datetime.now().strftime("%B %d, %Y"),
                df_items=summary
            )

            st.download_button(
                label="📊 Download DepEd ICS Form (Excel)",
                data=excel_data,
                file_name=f"ICS_{selected_school.replace(' ', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
        else:
            st.success("🎉 No pending dispatches for this school. All books have been received!")
    else:
        st.info("No dispatches logged for this school yet.")
